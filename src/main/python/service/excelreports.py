# vim: tabstop=8 expandtab shiftwidth=4 softtabstop=4
# -*- coding: utf-8 -*-
"""
:copyright: (c) 2020 Paolo Bernardi.
:license: GNU AGPL version 3, see LICENSE for more details.
"""

import calendar
from datetime import datetime, timedelta
import enum
import logging
import os
import shutil
from typing import cast, Dict, List

import openpyxl

from config import HelperType
from organization import Organization
from server.jira import Worklog
from service.public_holidays import PublicHolidays
from utils import change_locale, replace_path_vars


class ReportType(enum.Enum):
    forecast = 1
    final = 2
    monthly = 3


class ExcelReports:
    org: Organization
    helper: HelperType
    forecast_template: str
    forecast_report_dir: str
    final_template: str
    final_report_dir: str
    monthly_template: str
    monthly_report_dir: str
    suffix_forecast: str
    suffix_final: str
    prefix_monthly: str
    suffix_monthly: str
    max_days_back: int
    from_address: str
    rcpt_to_daily: List[str]
    rcpt_cc_daily: List[str]
    rcpt_to_monthly: List[str]
    rcpt_cc_monthly: List[str]
    first_log_row: int
    first_log_col: int
    sending_time_row: int
    sending_time_col: int
    ticket_replace: Dict[str, str]

    def __init__(self, org: Organization, helper: HelperType):
        self.org = org
        self.helper = helper
        params = cast(Dict[str, str], self.helper["parameters"])
        self.forecast_template = replace_path_vars(params["forecast_template"])
        self.forecast_report_dir = replace_path_vars(params["forecast_report_dir"])
        self.final_template = replace_path_vars(params["final_template"])
        self.final_report_dir = replace_path_vars(params["final_report_dir"])
        self.monthly_template = replace_path_vars(params["monthly_template"])
        self.monthly_report_dir = replace_path_vars(params["monthly_report_dir"])
        self.suffix_forecast = params["suffix_forecast"]
        self.suffix_final = params["suffix_final"]
        self.prefix_monthly = params["prefix_monthly"]
        self.suffix_monthly = params["suffix_monthly"]
        self.from_address = params["from_address"]
        self.rcpt_to_daily = params["rcpt_to_daily"].split(",")
        self.rcpt_cc_daily = params["rcpt_cc_daily"].split(",")
        self.rcpt_to_monthly = params["rcpt_to_monthly"].split(",")
        self.rcpt_cc_monthly = params["rcpt_cc_monthly"].split(",")
        self.max_days_back = int(params["max_days_back"])
        self.first_log_row = int(params["first_log_row"])
        self.first_log_col = int(params["first_log_col"])
        self.sending_time_row = int(params["sending_time_row"])
        self.sending_time_col = int(params["sending_time_col"])
        self.ticket_replace = dict(
            x.strip().split(":") for x in params["ticket_replace"].split(",")  # type: ignore
        )

    def get_nearest_time(self, reportType: ReportType, now: datetime):
        """
        Return the most usual entry/leave time for the specified
        report type, if applicable.
        """
        if reportType == ReportType.forecast:
            return now.replace(hour=9, minute=0, second=0)
        elif reportType == ReportType.final:
            return now.replace(hour=18, minute=0, second=0)
        else:
            return now

    def clone_fill(self, cell: openpyxl.cell.cell.Cell) -> openpyxl.styles.PatternFill:
        """
        Return a clone of the fill style for the specified cell.
        It currently clones only patternType and fgColor, as they are
        the parameters that I need here.
        """
        return openpyxl.styles.PatternFill(
            cell.fill.patternType, fgColor=cell.fill.fgColor
        )

    def init_report(self, report: str, reportType: ReportType, now: datetime):
        """
        Initialize the specified report file. Depending on the report type,
        date and time may be written and some cells may be blanked out, among
        other things.
        """
        now = self.get_nearest_time(reportType, now)
        wb = openpyxl.load_workbook(report)
        ws = wb.worksheets[0]  # All reports have 1 sheet only
        if reportType == ReportType.monthly:
            ws.cell(3, 6).value = now.strftime("%B").capitalize()
            ws.cell(3, 12).value = now.strftime("%Y")
            white_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF")
            weekend_fill = self.clone_fill(ws.cell(14, 8))
            holiday_fill = self.clone_fill(ws.cell(14, 17))
            public_holidays = PublicHolidays()
            for row in range(6, 12):
                for col in range(6, 37):
                    cell = ws.cell(row, col)
                    cell.value = ""
                    cell.fill = white_fill
            # Color weekends and holidays
            for day in range(1, calendar.monthrange(now.year, now.month)[1] + 1):
                cell = ws.cell(6, 5 + day).value = day
                new_date = now.replace(day=day)
                cell = ws.cell(7, 5 + day)
                if public_holidays.is_italian_public_holiday(
                    new_date
                ) or public_holidays.is_terni_saint_patron(new_date):
                    cell.fill = holiday_fill
                elif new_date.weekday() >= 5:  # 5 Saturday, 6 Sunday
                    cell.fill = weekend_fill
        elif reportType == ReportType.forecast:
            ws.cell(2, 3).value = now
            ws.cell(3, 3).value = now.strftime("%H.%M")
            for row in range(22, 38):
                for col in range(2, 7):
                    ws.cell(row, col).value = ""
        elif reportType == ReportType.final:
            ws.cell(2, 3).value = now
            ws.cell(3, 3).value = now.strftime("%H.%M")
        wb.save(report)

    def get_today_file(self, reportType: ReportType, now: datetime):
        # TODO: even though I've set the locale in the main function
        # I need to set it here as well, probably because it's another
        # thread or something like that...
        original_report_type = reportType
        original_now = now
        with change_locale("it_IT.utf8"):
            daysBack = 0
            originalFullPath = ""
            while daysBack < self.max_days_back:
                # This isn't a bug, it's a feature! I prefer to create a Forecast report based on the
                # Final report of the previous day.
                if daysBack == 1 and reportType == ReportType.forecast:
                    reportType = ReportType.final
                if reportType == ReportType.forecast:
                    file = "{}{}".format(now.strftime("%d%m%Y"), self.suffix_forecast)
                    subdir = now.strftime("%Y-%m")
                    fullPath = os.path.join(self.forecast_report_dir, subdir, file)
                elif reportType == ReportType.final:
                    file = "{}{}".format(now.strftime("%d%m%Y"), self.suffix_final)
                    subdir = now.strftime("%Y-%m")
                    fullPath = os.path.join(self.final_report_dir, subdir, file)
                elif reportType == ReportType.monthly:
                    file = "{}{}_{}{}".format(
                        self.prefix_monthly,
                        now.strftime("%B").capitalize(),
                        now.strftime("%Y"),
                        self.suffix_monthly,
                    )
                    subdir = now.strftime("%Y")
                    fullPath = os.path.join(self.monthly_report_dir, subdir, file)
                else:
                    return None
                if not originalFullPath:
                    originalFullPath = fullPath
                logging.debug(f"Attempting with {fullPath}")
                if fullPath and os.path.exists(fullPath):
                    break
                # This isn't a bug, it's a feature! I prefer to start create a Final report based on the
                # Forecast report of the current day. However, if the current day's Forecast report doesn't
                # exist, the past Final reports will be used.
                if daysBack == 0 and reportType == ReportType.final:
                    file = "{}{}".format(now.strftime("%d%m%Y"), self.suffix_forecast)
                    subdir = now.strftime("%Y-%m")
                    fullPath = os.path.join(self.forecast_report_dir, subdir, file)
                    if fullPath and os.path.exists(fullPath):
                        break
                daysBack += 1
                now -= timedelta(days=1)
            if not fullPath:
                if reportType == ReportType.forecast:
                    fullPath = self.forecast_template
                elif reportType == ReportType.final:
                    fullPath = self.final_template
                elif reportType == ReportType.monthly:
                    fullPath = self.monthly_template
                else:
                    return None
            if originalFullPath != fullPath:
                destDir = os.path.dirname(originalFullPath)
                if not os.path.exists(destDir):
                    os.makedirs(destDir)
                shutil.copyfile(fullPath, originalFullPath)
                self.init_report(originalFullPath, original_report_type, original_now)
            return originalFullPath

    def get_sending_time(self, now: datetime):
        xls = self.get_today_file(ReportType.forecast, now)
        workbook = openpyxl.load_workbook(xls)
        sheet = workbook.worksheets[0]
        str_sending_time = sheet.cell(
            self.sending_time_row, self.sending_time_col
        ).value
        sending_time = datetime.strptime(str_sending_time, "%H.%M")
        return sending_time.replace(
            year=now.year, month=now.month, day=now.day, second=0, microsecond=0
        )

    def parse_worklogs(
        self, now: datetime, sending_time: datetime, jira_user: str
    ) -> List[Worklog]:
        def parse_row(sheet, row, start_time):
            description = sheet.cell(row, self.first_log_col).value
            if not description or not description.strip():
                return None
            ticket = description[: description.find(":")]
            duration = sheet.cell(row, self.first_log_col + 3).value
            return Worklog(ticket, jira_user, description, start_time, duration)

        xls = self.get_today_file(ReportType.final, now)
        workbook = openpyxl.load_workbook(xls)
        sheet = workbook.worksheets[0]
        row = self.first_log_row
        lunch_break = False
        worklogs = []
        start_time = sending_time
        data = parse_row(sheet, row, start_time)
        while data:
            for wrong, correct in self.ticket_replace.items():
                data.ticket = data.ticket.replace(wrong, correct)
                data.description = data.description.replace(wrong, correct)
            end_time = start_time + timedelta(hours=data.duration)
            worklogs.append(data)
            row += 1
            start_time = end_time
            if start_time.hour >= 13 and not lunch_break:
                start_time += timedelta(hours=1)
                lunch_break = True
            data = parse_row(sheet, row, start_time)
        return worklogs
